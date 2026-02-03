#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
広告事業取引監査用ダミーデータ生成スクリプト
受発注請求データExcelの生成
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# テーマ設定
THEME = {
    'primary': '1F4E79',
    'light': 'D6E3F0',
    'accent': '1F4E79',
}

SERIF_FONT = 'Source Serif Pro'
SANS_FONT = 'Source Sans Pro'

# ボーダー定義
outer_border = Side(style='thin', color='D1D1D1')
header_bottom = Side(style='medium', color=THEME['primary'])
inner_horizontal = Side(style='thin', color='D1D1D1')
no_border = Side(style=None)

def apply_data_block_borders(ws, start_row, end_row, start_col, end_col, has_header=True):
    """データブロックにボーダーを適用"""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            left = outer_border if col == start_col else no_border
            right = outer_border if col == end_col else no_border
            top = outer_border if row == start_row else inner_horizontal
            if has_header and row == start_row:
                bottom = header_bottom
            elif row == end_row:
                bottom = outer_border
            else:
                bottom = inner_horizontal
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

def create_order_invoice_data():
    """受発注請求データを作成"""
    wb = Workbook()
    
    # === 概要シート ===
    ws_overview = wb.active
    ws_overview.title = "概要"
    ws_overview.sheet_view.showGridLines = False
    ws_overview.column_dimensions['A'].width = 3
    
    # タイトル
    ws_overview['B2'] = "受発注請求データ一覧"
    ws_overview['B2'].font = Font(name=SERIF_FONT, size=20, bold=True, color=THEME['primary'])
    
    ws_overview['B3'] = "広告事業取引監査用テストデータ"
    ws_overview['B3'].font = Font(name=SANS_FONT, size=12, color='666666')
    
    ws_overview['B5'] = f"生成日: {datetime.now().strftime('%Y年%m月%d日')}"
    ws_overview['B5'].font = Font(name=SANS_FONT, size=10, italic=True, color='666666')
    
    # 概要情報
    ws_overview['B7'] = "データ概要"
    ws_overview['B7'].font = Font(name=SERIF_FONT, size=14, bold=True, color=THEME['primary'])
    
    summary_data = [
        ["総取引件数", "10件"],
        ["正常取引", "3件"],
        ["検知対象取引", "7件"],
        ["データ期間", "2025年6月〜2026年1月"],
    ]
    
    for i, (label, value) in enumerate(summary_data, start=8):
        ws_overview.cell(row=i, column=2, value=label).font = Font(name=SANS_FONT, size=11)
        ws_overview.cell(row=i, column=3, value=value).font = Font(name=SANS_FONT, size=11, bold=True)
    
    # シート目次
    ws_overview['B14'] = "シート構成"
    ws_overview['B14'].font = Font(name=SERIF_FONT, size=14, bold=True, color=THEME['primary'])
    
    sheets_info = [
        ["取引一覧", "全取引の基本情報"],
        ["見積データ", "見積書から抽出したデータ"],
        ["発注データ", "発注書から抽出したデータ"],
        ["請求データ", "請求書から抽出したデータ"],
        ["検収データ", "検収情報"],
        ["不整合チェック", "自動検出された不整合"],
    ]
    
    for i, (sheet_name, desc) in enumerate(sheets_info, start=15):
        cell = ws_overview.cell(row=i, column=2, value=sheet_name)
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.font = Font(name=SANS_FONT, size=11, color=THEME['accent'], underline='single')
        ws_overview.cell(row=i, column=3, value=desc).font = Font(name=SANS_FONT, size=11, color='666666')
    
    ws_overview.column_dimensions['B'].width = 20
    ws_overview.column_dimensions['C'].width = 40
    
    # === 取引一覧シート ===
    ws_list = wb.create_sheet("取引一覧")
    ws_list.sheet_view.showGridLines = False
    ws_list.column_dimensions['A'].width = 3
    
    ws_list['B2'] = "取引一覧"
    ws_list['B2'].font = Font(name=SERIF_FONT, size=16, bold=True, color=THEME['primary'])
    
    headers = ["取引ID", "取引名", "発注先", "発注先業種", "見積金額", "発注金額", "請求金額", "ステータス", "不整合フラグ"]
    
    transactions = [
        ["TX-001", "Web広告クリエイティブ制作", "株式会社クリエイティブワークス", "デザイン制作", 500000, 500000, 500000, "完了", ""],
        ["TX-002", "SNS運用代行業務（3ヶ月）", "デジタルマーケティング合同会社", "マーケティング", 900000, 900000, 900000, "完了", ""],
        ["TX-003", "動画広告制作（15秒CM）", "映像制作株式会社メディアプロ", "映像制作", 1200000, 1200000, 1200000, "完了", ""],
        ["TX-101", "ランディングページ制作", "株式会社クリエイティブワークス", "デザイン制作", 500000, 480000, 450000, "完了", "金額不整合"],
        ["TX-102", "リスティング広告運用代行（3ヶ月）", "個人事業主 山田太郎", "広告運用", 1500000, 1500000, 1500000, "進行中", "日付不整合"],
        ["TX-103", "広告効果分析・市場調査レポート作成", "株式会社建設コンサルタント", "建設コンサル", 800000, 800000, 800000, "完了", "発注先属性不整合"],
        ["TX-104", "SEOコンテンツ記事作成", "ライティングプロ株式会社", "ライティング", 5000000, 5000000, 5000000, "完了", "金額と内容不整合"],
        ["TX-105", "ブランディング戦略策定", "株式会社ストラテジーパートナーズ", "コンサルティング", 2000000, 2000000, 2000000, "完了", "検収物独自性欠如"],
        ["TX-106", "インフルエンサーマーケティング施策", "インフルエンスマーケティング株式会社", "インフルエンサー", 2400000, 3000000, 2100000, "完了", "数量・金額不整合"],
        ["TX-107", "広告効果測定ダッシュボード構築", "システム開発株式会社テクノソリューション", "システム開発", 4500000, 4500000, 4500000, "完了", "日付不整合"],
    ]
    
    # ヘッダー
    header_fill = PatternFill(start_color=THEME['primary'], end_color=THEME['primary'], fill_type='solid')
    header_font = Font(name=SERIF_FONT, size=10, bold=True, color='FFFFFF')
    
    for col, header in enumerate(headers, start=2):
        cell = ws_list.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # データ
    data_font = Font(name=SANS_FONT, size=10)
    warning_fill = PatternFill(start_color='FFCCBC', end_color='FFCCBC', fill_type='solid')
    
    for row_idx, tx in enumerate(transactions, start=5):
        for col_idx, value in enumerate(tx, start=2):
            cell = ws_list.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            if col_idx in [6, 7, 8]:  # 金額列
                cell.number_format = '¥#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            
            # 不整合がある行をハイライト
            if tx[-1]:  # 不整合フラグがある場合
                cell.fill = warning_fill
    
    apply_data_block_borders(ws_list, 4, 4 + len(transactions), 2, 2 + len(headers) - 1)
    
    # 列幅設定
    col_widths = [12, 35, 35, 18, 15, 15, 15, 12, 20]
    for i, width in enumerate(col_widths, start=2):
        ws_list.column_dimensions[get_column_letter(i)].width = width
    
    # === 見積データシート ===
    ws_estimate = wb.create_sheet("見積データ")
    ws_estimate.sheet_view.showGridLines = False
    ws_estimate.column_dimensions['A'].width = 3
    
    ws_estimate['B2'] = "見積データ"
    ws_estimate['B2'].font = Font(name=SERIF_FONT, size=16, bold=True, color=THEME['primary'])
    
    est_headers = ["取引ID", "見積番号", "見積日", "有効期限", "発注先", "件名", "金額（税抜）", "金額（税込）", "明細数量", "単位"]
    
    estimate_data = [
        ["TX-001", "EST-001", "2025/07/25", "2025/08/25", "株式会社クリエイティブワークス", "Web広告クリエイティブ制作", 500000, 550000, 10, "点"],
        ["TX-002", "EST-002", "2025/06/20", "2025/07/20", "デジタルマーケティング合同会社", "SNS運用代行業務（3ヶ月）", 900000, 990000, 3, "ヶ月"],
        ["TX-003", "EST-003", "2025/06/10", "2025/07/10", "映像制作株式会社メディアプロ", "動画広告制作（15秒CM）", 1200000, 1320000, 1, "式"],
        ["TX-101", "EST-101", "2025/08/25", "2025/09/25", "株式会社クリエイティブワークス", "ランディングページ制作", 500000, 550000, 1, "式"],
        ["TX-102", "EST-102", "2025/10/15", "2025/11/15", "個人事業主 山田太郎", "リスティング広告運用代行（3ヶ月）", 1500000, 1650000, 3, "ヶ月"],
        ["TX-103", "EST-103", "2025/08/10", "2025/09/10", "株式会社建設コンサルタント", "広告効果分析・市場調査レポート作成", 800000, 880000, 1, "式"],
        ["TX-104", "EST-104", "2025/07/15", "2025/08/15", "ライティングプロ株式会社", "SEOコンテンツ記事作成", 5000000, 5500000, 50, "本"],
        ["TX-105", "EST-105", "2025/09/05", "2025/10/05", "株式会社ストラテジーパートナーズ", "ブランディング戦略策定", 2000000, 2200000, 1, "式"],
        ["TX-106", "EST-106", "2025/07/25", "2025/08/25", "インフルエンスマーケティング株式会社", "インフルエンサーマーケティング施策", 2400000, 2640000, 8, "名"],
        ["TX-107", "EST-107", "2025/06/25", "2025/07/25", "システム開発株式会社テクノソリューション", "広告効果測定ダッシュボード構築", 4500000, 4950000, 1, "式"],
    ]
    
    for col, header in enumerate(est_headers, start=2):
        cell = ws_estimate.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_idx, data in enumerate(estimate_data, start=5):
        for col_idx, value in enumerate(data, start=2):
            cell = ws_estimate.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            if col_idx in [8, 9]:  # 金額列
                cell.number_format = '¥#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    
    apply_data_block_borders(ws_estimate, 4, 4 + len(estimate_data), 2, 2 + len(est_headers) - 1)
    
    est_col_widths = [12, 12, 12, 12, 35, 35, 15, 15, 12, 8]
    for i, width in enumerate(est_col_widths, start=2):
        ws_estimate.column_dimensions[get_column_letter(i)].width = width
    
    # === 発注データシート ===
    ws_order = wb.create_sheet("発注データ")
    ws_order.sheet_view.showGridLines = False
    ws_order.column_dimensions['A'].width = 3
    
    ws_order['B2'] = "発注データ"
    ws_order['B2'].font = Font(name=SERIF_FONT, size=16, bold=True, color=THEME['primary'])
    
    ord_headers = ["取引ID", "発注番号", "発注日", "納期", "発注先", "件名", "金額（税抜）", "金額（税込）", "明細数量", "単位", "発注者"]
    
    order_data = [
        ["TX-001", "PO-001", "2025/08/01", "2025/08/31", "株式会社クリエイティブワークス", "Web広告クリエイティブ制作", 500000, 550000, 10, "点", "田中 一郎"],
        ["TX-002", "PO-002", "2025/07/01", "2025/09/30", "デジタルマーケティング合同会社", "SNS運用代行業務（3ヶ月）", 900000, 990000, 3, "ヶ月", "山本 美咲"],
        ["TX-003", "PO-003", "2025/06/15", "2025/07/31", "映像制作株式会社メディアプロ", "動画広告制作（15秒CM）", 1200000, 1320000, 1, "式", "伊藤 直樹"],
        ["TX-101", "PO-101", "2025/09/01", "2025/09/30", "株式会社クリエイティブワークス", "ランディングページ制作", 480000, 528000, 1, "式", "田中 一郎"],
        ["TX-102", "PO-102", "2025/10/01", "2025/12/31", "個人事業主 山田太郎", "リスティング広告運用代行（3ヶ月）", 1500000, 1650000, 3, "ヶ月", "中村 雅人"],
        ["TX-103", "PO-103", "2025/08/15", "2025/09/15", "株式会社建設コンサルタント", "広告効果分析・市場調査レポート作成", 800000, 880000, 1, "式", "小林 裕子"],
        ["TX-104", "PO-104", "2025/07/20", "2025/08/31", "ライティングプロ株式会社", "SEOコンテンツ記事作成", 5000000, 5500000, 50, "本", "吉田 浩二"],
        ["TX-105", "PO-105", "2025/09/10", "2025/10/31", "株式会社ストラテジーパートナーズ", "ブランディング戦略策定", 2000000, 2200000, 1, "式", "井上 真理"],
        ["TX-106", "PO-106", "2025/08/01", "2025/09/30", "インフルエンスマーケティング株式会社", "インフルエンサーマーケティング施策", 3000000, 3300000, 10, "名", "斎藤 恵"],
        ["TX-107", "PO-107", "2025/07/01", "2025/10/15", "システム開発株式会社テクノソリューション", "広告効果測定ダッシュボード構築", 4500000, 4950000, 1, "式", "森田 優子"],
    ]
    
    for col, header in enumerate(ord_headers, start=2):
        cell = ws_order.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_idx, data in enumerate(order_data, start=5):
        for col_idx, value in enumerate(data, start=2):
            cell = ws_order.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            if col_idx in [8, 9]:  # 金額列
                cell.number_format = '¥#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    
    apply_data_block_borders(ws_order, 4, 4 + len(order_data), 2, 2 + len(ord_headers) - 1)
    
    ord_col_widths = [12, 12, 12, 12, 35, 35, 15, 15, 12, 8, 15]
    for i, width in enumerate(ord_col_widths, start=2):
        ws_order.column_dimensions[get_column_letter(i)].width = width
    
    # === 請求データシート ===
    ws_invoice = wb.create_sheet("請求データ")
    ws_invoice.sheet_view.showGridLines = False
    ws_invoice.column_dimensions['A'].width = 3
    
    ws_invoice['B2'] = "請求データ"
    ws_invoice['B2'].font = Font(name=SERIF_FONT, size=16, bold=True, color=THEME['primary'])
    
    inv_headers = ["取引ID", "請求番号", "請求日", "支払期限", "発注先", "件名", "金額（税抜）", "金額（税込）", "支払状況"]
    
    invoice_data = [
        ["TX-001", "INV-001", "2025/09/05", "2025/10/31", "株式会社クリエイティブワークス", "Web広告クリエイティブ制作", 500000, 550000, "支払済"],
        ["TX-002", "INV-002", "2025/10/05", "2025/11/30", "デジタルマーケティング合同会社", "SNS運用代行業務（3ヶ月）", 900000, 990000, "支払済"],
        ["TX-003", "INV-003", "2025/08/05", "2025/09/30", "映像制作株式会社メディアプロ", "動画広告制作（15秒CM）", 1200000, 1320000, "支払済"],
        ["TX-101", "INV-101", "2025/10/05", "2025/11/30", "株式会社クリエイティブワークス", "ランディングページ制作", 450000, 495000, "支払済"],
        ["TX-102", "INV-102", "2026/01/10", "2026/02/28", "個人事業主 山田太郎", "リスティング広告運用代行（3ヶ月）", 1500000, 1650000, "未払"],
        ["TX-103", "INV-103", "2025/09/20", "2025/10/31", "株式会社建設コンサルタント", "広告効果分析・市場調査レポート作成", 800000, 880000, "支払済"],
        ["TX-104", "INV-104", "2025/09/05", "2025/10/31", "ライティングプロ株式会社", "SEOコンテンツ記事作成", 5000000, 5500000, "支払済"],
        ["TX-105", "INV-105", "2025/11/05", "2025/12/31", "株式会社ストラテジーパートナーズ", "ブランディング戦略策定", 2000000, 2200000, "支払済"],
        ["TX-106", "INV-106", "2025/10/05", "2025/11/30", "インフルエンスマーケティング株式会社", "インフルエンサーマーケティング施策", 2100000, 2310000, "支払済"],
        ["TX-107", "INV-107", "2025/09/30", "2025/10/31", "システム開発株式会社テクノソリューション", "広告効果測定ダッシュボード構築", 4500000, 4950000, "支払済"],
    ]
    
    for col, header in enumerate(inv_headers, start=2):
        cell = ws_invoice.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_idx, data in enumerate(invoice_data, start=5):
        for col_idx, value in enumerate(data, start=2):
            cell = ws_invoice.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            if col_idx in [8, 9]:  # 金額列
                cell.number_format = '¥#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    
    apply_data_block_borders(ws_invoice, 4, 4 + len(invoice_data), 2, 2 + len(inv_headers) - 1)
    
    inv_col_widths = [12, 12, 12, 12, 35, 35, 15, 15, 12]
    for i, width in enumerate(inv_col_widths, start=2):
        ws_invoice.column_dimensions[get_column_letter(i)].width = width
    
    # === 検収データシート ===
    ws_acceptance = wb.create_sheet("検収データ")
    ws_acceptance.sheet_view.showGridLines = False
    ws_acceptance.column_dimensions['A'].width = 3
    
    ws_acceptance['B2'] = "検収データ"
    ws_acceptance['B2'].font = Font(name=SERIF_FONT, size=16, bold=True, color=THEME['primary'])
    
    acc_headers = ["取引ID", "検収番号", "納品日", "検収日", "検収担当者", "検収物種別", "検収物ファイル", "検収数量", "検収結果", "備考"]
    
    acceptance_data = [
        ["TX-001", "ACC-001", "2025/08/31", "2025/09/05", "田中 一郎", "画像", "TX-001_バナー広告.zip", 10, "合格", "オリジナルデザイン10点"],
        ["TX-002", "ACC-002", "2025/09/30", "2025/10/05", "山本 美咲", "Excel", "TX-002_運用レポート.xlsx", 3, "合格", "月次レポート3ヶ月分"],
        ["TX-003", "ACC-003", "2025/07/31", "2025/08/05", "伊藤 直樹", "PDF", "TX-003_動画仕様書.pdf", 1, "合格", "15秒CM動画1本"],
        ["TX-101", "ACC-101", "2025/09/30", "2025/10/05", "田中 一郎", "PDF", "TX-101_LP画面.pdf", 1, "合格", "LP1ページ"],
        ["TX-102", "ACC-102", "2025/12/31", "2026/01/10", "中村 雅人", "Excel", "TX-102_広告レポート.xlsx", 3, "合格", "週次レポート"],
        ["TX-103", "ACC-103", "2025/09/15", "2025/09/20", "小林 裕子", "PDF", "TX-103_市場調査.pdf", 1, "合格", "調査レポート"],
        ["TX-104", "ACC-104", "2025/08/31", "2025/09/05", "吉田 浩二", "Excel", "TX-104_記事一覧.xlsx", 50, "合格", "SEO記事50本"],
        ["TX-105", "ACC-105", "2025/10/31", "2025/11/05", "井上 真理", "PowerPoint", "TX-105_戦略資料.pptx", 1, "合格", "汎用テンプレート使用"],
        ["TX-106", "ACC-106", "2025/09/30", "2025/10/05", "斎藤 恵", "Excel", "TX-106_投稿一覧.xlsx", 7, "合格", "投稿7件のみ"],
        ["TX-107", "ACC-107", "2025/10/15", "2025/09/30", "森田 優子", "PDF", "TX-107_システム仕様書.pdf", 1, "合格", ""],
    ]
    
    for col, header in enumerate(acc_headers, start=2):
        cell = ws_acceptance.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_idx, data in enumerate(acceptance_data, start=5):
        for col_idx, value in enumerate(data, start=2):
            cell = ws_acceptance.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    
    apply_data_block_borders(ws_acceptance, 4, 4 + len(acceptance_data), 2, 2 + len(acc_headers) - 1)
    
    acc_col_widths = [12, 12, 12, 12, 15, 15, 30, 12, 10, 25]
    for i, width in enumerate(acc_col_widths, start=2):
        ws_acceptance.column_dimensions[get_column_letter(i)].width = width
    
    # === 不整合チェックシート ===
    ws_check = wb.create_sheet("不整合チェック")
    ws_check.sheet_view.showGridLines = False
    ws_check.column_dimensions['A'].width = 3
    
    ws_check['B2'] = "不整合チェック結果"
    ws_check['B2'].font = Font(name=SERIF_FONT, size=16, bold=True, color=THEME['primary'])
    
    ws_check['B3'] = "自動検出された不整合一覧"
    ws_check['B3'].font = Font(name=SANS_FONT, size=11, color='666666')
    
    chk_headers = ["取引ID", "不整合種別", "検出内容", "見積値", "発注値", "請求/検収値", "差異", "重要度"]
    
    check_data = [
        ["TX-101", "金額不整合", "見積・発注・請求金額が一致しない", "¥500,000", "¥480,000", "¥450,000", "¥50,000", "高"],
        ["TX-102", "日付不整合", "見積日が発注日より後", "2025/10/15", "2025/10/01", "-", "14日", "高"],
        ["TX-103", "発注先属性不整合", "建設コンサル会社に広告分析を発注", "建設コンサル", "広告効果分析", "-", "-", "中"],
        ["TX-104", "金額と内容不整合", "SEO記事1本あたり10万円は相場の5-10倍", "50本", "¥5,000,000", "¥100,000/本", "相場比+400%", "高"],
        ["TX-105", "検収物独自性欠如", "成果物が汎用テンプレートのみ", "-", "-", "テンプレート", "-", "中"],
        ["TX-106", "数量不整合", "見積・発注・検収の数量が一致しない", "8名", "10名", "7名", "3名", "高"],
        ["TX-106", "金額不整合", "数量変更に伴う金額不整合", "¥2,400,000", "¥3,000,000", "¥2,100,000", "¥900,000", "高"],
        ["TX-107", "日付不整合", "検収日が納品日より前", "納品:2025/10/15", "検収:2025/09/30", "-", "-15日", "高"],
    ]
    
    for col, header in enumerate(chk_headers, start=2):
        cell = ws_check.cell(row=5, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    high_fill = PatternFill(start_color='FFCCBC', end_color='FFCCBC', fill_type='solid')
    medium_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    
    for row_idx, data in enumerate(check_data, start=6):
        for col_idx, value in enumerate(data, start=2):
            cell = ws_check.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            
            # 重要度に応じてハイライト
            if data[-1] == "高":
                cell.fill = high_fill
            elif data[-1] == "中":
                cell.fill = medium_fill
    
    apply_data_block_borders(ws_check, 5, 5 + len(check_data), 2, 2 + len(chk_headers) - 1)
    
    chk_col_widths = [12, 20, 40, 18, 18, 18, 15, 10]
    for i, width in enumerate(chk_col_widths, start=2):
        ws_check.column_dimensions[get_column_letter(i)].width = width
    
    # 保存
    output_path = '/home/ubuntu/ad_audit_test_data/受発注請求データ/受発注請求一覧.xlsx'
    wb.save(output_path)
    print(f"受発注請求データExcel生成完了: {output_path}")

if __name__ == "__main__":
    create_order_invoice_data()
