#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
広告事業取引監査用ダミーデータ生成スクリプト
検収用資料（PDF、Excel）の生成
"""

from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
import os

# テーマ設定
THEME = {
    'primary': '1F4E79',
    'light': 'D6E3F0',
}

outer_border = Side(style='thin', color='D1D1D1')
header_bottom = Side(style='medium', color=THEME['primary'])
inner_horizontal = Side(style='thin', color='D1D1D1')
no_border = Side(style=None)

def apply_borders(ws, start_row, end_row, start_col, end_col, has_header=True):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            left = outer_border if col == start_col else no_border
            right = outer_border if col == end_col else no_border
            top = outer_border if row == start_row else inner_horizontal
            if has_header and row == start_row:
                bottom = header_bottom
            elif row == end_row:
                bottom = outer_border
            else:
                bottom = inner_horizontal
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def create_tx001_banner_report():
    """TX-001: バナー広告制作完了報告書（PDF）"""
    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page { size: A4; margin: 20mm; }
            body { font-family: 'Noto Sans CJK JP', sans-serif; font-size: 10pt; line-height: 1.6; }
            .title { text-align: center; font-size: 18pt; font-weight: bold; margin-bottom: 20px; }
            .info-table { width: 100%; margin: 20px 0; border-collapse: collapse; }
            .info-table td { padding: 8px; border: 1px solid #ddd; }
            .info-table .label { background: #f5f5f5; width: 30%; font-weight: bold; }
            .deliverables { margin: 20px 0; }
            .deliverables h3 { color: #1F4E79; border-bottom: 2px solid #1F4E79; padding-bottom: 5px; }
            .file-list { margin: 10px 0; }
            .file-list li { margin: 5px 0; }
            .signature { margin-top: 40px; text-align: right; }
        </style>
    </head>
    <body>
        <div class="title">納品完了報告書</div>
        
        <table class="info-table">
            <tr><td class="label">取引ID</td><td>TX-001</td></tr>
            <tr><td class="label">件名</td><td>Web広告クリエイティブ制作</td></tr>
            <tr><td class="label">発注元</td><td>株式会社サンプル広告</td></tr>
            <tr><td class="label">納品先担当</td><td>田中 一郎 様</td></tr>
            <tr><td class="label">納品日</td><td>2025年8月31日</td></tr>
            <tr><td class="label">納品者</td><td>株式会社クリエイティブワークス 佐藤 花子</td></tr>
        </table>
        
        <div class="deliverables">
            <h3>納品物一覧</h3>
            <p>ご発注いただきましたバナー広告デザイン10点を納品いたします。</p>
            
            <table class="info-table">
                <tr style="background: #1F4E79; color: white;">
                    <th>No.</th><th>ファイル名</th><th>サイズ</th><th>用途</th>
                </tr>
                <tr><td>1</td><td>banner_300x250_01.png</td><td>300×250px</td><td>レクタングル広告</td></tr>
                <tr><td>2</td><td>banner_300x250_02.png</td><td>300×250px</td><td>レクタングル広告</td></tr>
                <tr><td>3</td><td>banner_728x90_01.png</td><td>728×90px</td><td>リーダーボード</td></tr>
                <tr><td>4</td><td>banner_728x90_02.png</td><td>728×90px</td><td>リーダーボード</td></tr>
                <tr><td>5</td><td>banner_160x600_01.png</td><td>160×600px</td><td>スカイスクレイパー</td></tr>
                <tr><td>6</td><td>banner_160x600_02.png</td><td>160×600px</td><td>スカイスクレイパー</td></tr>
                <tr><td>7</td><td>banner_320x50_01.png</td><td>320×50px</td><td>モバイルバナー</td></tr>
                <tr><td>8</td><td>banner_320x50_02.png</td><td>320×50px</td><td>モバイルバナー</td></tr>
                <tr><td>9</td><td>banner_300x600_01.png</td><td>300×600px</td><td>ハーフページ</td></tr>
                <tr><td>10</td><td>banner_300x600_02.png</td><td>300×600px</td><td>ハーフページ</td></tr>
            </table>
            
            <h3>デザインコンセプト</h3>
            <p>「信頼と革新」をテーマに、貴社のブランドカラーであるブルーを基調としたデザインを制作いたしました。
            ターゲット層である30-40代ビジネスパーソンに訴求するため、シンプルかつ洗練されたビジュアルを心がけました。</p>
            
            <h3>備考</h3>
            <ul>
                <li>全ファイルはPNG形式（透過対応）で納品</li>
                <li>AI形式の元データは別途ご要望があればお送りします</li>
                <li>修正対応は納品後2週間以内、2回まで無償対応</li>
            </ul>
        </div>
        
        <div class="signature">
            <p>以上、ご確認のほどよろしくお願いいたします。</p>
            <p>株式会社クリエイティブワークス<br>
            デザイン部 佐藤 花子</p>
        </div>
    </body>
    </html>
    """
    output_path = '/home/ubuntu/ad_audit_test_data/検収資料/PDF/TX-001_バナー広告納品報告書.pdf'
    HTML(string=html).write_pdf(output_path)
    print(f"生成完了: {output_path}")


def create_tx002_sns_report():
    """TX-002: SNS運用レポート（Excel）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "運用サマリー"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    
    # タイトル
    ws['B2'] = "SNS運用代行 月次レポート"
    ws['B2'].font = Font(name='Source Serif Pro', size=18, bold=True, color=THEME['primary'])
    
    ws['B3'] = "期間: 2025年7月〜9月（3ヶ月）"
    ws['B3'].font = Font(name='Source Sans Pro', size=11, color='666666')
    
    # サマリー
    ws['B5'] = "運用実績サマリー"
    ws['B5'].font = Font(name='Source Serif Pro', size=14, bold=True, color=THEME['primary'])
    
    header_fill = PatternFill(start_color=THEME['primary'], end_color=THEME['primary'], fill_type='solid')
    header_font = Font(name='Source Serif Pro', size=10, bold=True, color='FFFFFF')
    data_font = Font(name='Source Sans Pro', size=10)
    
    headers = ["指標", "7月", "8月", "9月", "合計/平均"]
    for col, h in enumerate(headers, start=2):
        cell = ws.cell(row=7, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    data = [
        ["Instagram投稿数", 22, 21, 23, 66],
        ["Instagram フォロワー増加", 1250, 1480, 1320, 4050],
        ["Instagram エンゲージメント率", "3.2%", "3.5%", "3.8%", "3.5%"],
        ["Twitter投稿数", 45, 42, 48, 135],
        ["Twitter フォロワー増加", 890, 1020, 950, 2860],
        ["Twitter エンゲージメント率", "2.1%", "2.4%", "2.6%", "2.4%"],
    ]
    
    for row_idx, row_data in enumerate(data, start=8):
        for col_idx, value in enumerate(row_data, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            if col_idx == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            else:
                cell.alignment = Alignment(horizontal='right', vertical='center')
    
    apply_borders(ws, 7, 13, 2, 6)
    
    for i, w in enumerate([25, 12, 12, 12, 15], start=2):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    # 月別詳細シート
    ws2 = wb.create_sheet("投稿詳細")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions['A'].width = 3
    
    ws2['B2'] = "投稿詳細一覧（抜粋）"
    ws2['B2'].font = Font(name='Source Serif Pro', size=16, bold=True, color=THEME['primary'])
    
    detail_headers = ["日付", "プラットフォーム", "投稿内容", "いいね", "コメント", "シェア"]
    for col, h in enumerate(detail_headers, start=2):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    detail_data = [
        ["2025/07/05", "Instagram", "新商品紹介キャンペーン告知", 342, 28, 15],
        ["2025/07/12", "Instagram", "お客様の声紹介", 289, 45, 22],
        ["2025/07/20", "Twitter", "業界ニュースシェア", 156, 12, 89],
        ["2025/08/03", "Instagram", "夏季限定商品PR", 512, 67, 34],
        ["2025/08/15", "Twitter", "キャンペーン告知", 234, 18, 156],
        ["2025/09/01", "Instagram", "秋の新作発表", 478, 52, 28],
        ["2025/09/18", "Twitter", "イベント告知", 198, 23, 112],
        ["2025/09/25", "Instagram", "お客様インタビュー", 356, 41, 19],
    ]
    
    for row_idx, row_data in enumerate(detail_data, start=5):
        for col_idx, value in enumerate(row_data, start=2):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left' if col_idx <= 4 else 'right', vertical='center')
    
    apply_borders(ws2, 4, 12, 2, 7)
    
    for i, w in enumerate([12, 15, 35, 10, 10, 10], start=2):
        ws2.column_dimensions[get_column_letter(i)].width = w
    
    output_path = '/home/ubuntu/ad_audit_test_data/検収資料/Excel/TX-002_SNS運用レポート.xlsx'
    wb.save(output_path)
    print(f"生成完了: {output_path}")


def create_tx003_video_spec():
    """TX-003: 動画広告仕様書（PDF）"""
    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page { size: A4; margin: 20mm; }
            body { font-family: 'Noto Sans CJK JP', sans-serif; font-size: 10pt; line-height: 1.6; }
            .title { text-align: center; font-size: 18pt; font-weight: bold; margin-bottom: 20px; color: #1F4E79; }
            .section { margin: 20px 0; }
            .section h2 { color: #1F4E79; border-left: 4px solid #1F4E79; padding-left: 10px; font-size: 14pt; }
            table { width: 100%; border-collapse: collapse; margin: 15px 0; }
            th, td { padding: 8px; border: 1px solid #ddd; text-align: left; }
            th { background: #1F4E79; color: white; }
            .timeline { background: #f9f9f9; padding: 15px; border-radius: 5px; }
            .timeline-item { margin: 10px 0; padding-left: 20px; border-left: 3px solid #1F4E79; }
        </style>
    </head>
    <body>
        <div class="title">動画広告制作 納品仕様書</div>
        
        <div class="section">
            <h2>基本情報</h2>
            <table>
                <tr><th width="30%">項目</th><th>内容</th></tr>
                <tr><td>取引ID</td><td>TX-003</td></tr>
                <tr><td>制作物</td><td>15秒テレビCM動画</td></tr>
                <tr><td>クライアント</td><td>株式会社サンプル広告</td></tr>
                <tr><td>制作会社</td><td>映像制作株式会社メディアプロ</td></tr>
                <tr><td>納品日</td><td>2025年7月31日</td></tr>
            </table>
        </div>
        
        <div class="section">
            <h2>動画仕様</h2>
            <table>
                <tr><th width="30%">項目</th><th>仕様</th></tr>
                <tr><td>尺</td><td>15秒</td></tr>
                <tr><td>解像度</td><td>1920×1080 (Full HD)</td></tr>
                <tr><td>フレームレート</td><td>29.97fps</td></tr>
                <tr><td>コーデック</td><td>H.264 / ProRes 422</td></tr>
                <tr><td>音声</td><td>ステレオ 48kHz / 16bit</td></tr>
                <tr><td>納品形式</td><td>MP4, MOV</td></tr>
            </table>
        </div>
        
        <div class="section">
            <h2>タイムライン構成</h2>
            <div class="timeline">
                <div class="timeline-item">
                    <strong>0:00-0:03</strong> オープニング<br>
                    ロゴアニメーション、キャッチコピー表示
                </div>
                <div class="timeline-item">
                    <strong>0:03-0:10</strong> メインパート<br>
                    商品・サービス紹介映像、ナレーション
                </div>
                <div class="timeline-item">
                    <strong>0:10-0:13</strong> ベネフィット訴求<br>
                    主要メリット3点をテロップ表示
                </div>
                <div class="timeline-item">
                    <strong>0:13-0:15</strong> エンディング<br>
                    CTA、企業ロゴ、URL表示
                </div>
            </div>
        </div>
        
        <div class="section">
            <h2>納品ファイル一覧</h2>
            <table>
                <tr><th>ファイル名</th><th>形式</th><th>用途</th></tr>
                <tr><td>TX003_CM_15sec_master.mov</td><td>ProRes 422</td><td>マスターデータ</td></tr>
                <tr><td>TX003_CM_15sec_web.mp4</td><td>H.264</td><td>Web配信用</td></tr>
                <tr><td>TX003_CM_15sec_preview.mp4</td><td>H.264 (低ビットレート)</td><td>プレビュー用</td></tr>
            </table>
        </div>
        
        <div class="section">
            <h2>制作クレジット</h2>
            <table>
                <tr><th>役割</th><th>担当</th></tr>
                <tr><td>プロデューサー</td><td>高橋 健一</td></tr>
                <tr><td>ディレクター</td><td>山田 誠</td></tr>
                <tr><td>撮影</td><td>佐々木 隆</td></tr>
                <tr><td>編集</td><td>田村 美香</td></tr>
                <tr><td>ナレーション</td><td>鈴木 花子</td></tr>
            </table>
        </div>
    </body>
    </html>
    """
    output_path = '/home/ubuntu/ad_audit_test_data/検収資料/PDF/TX-003_動画広告仕様書.pdf'
    HTML(string=html).write_pdf(output_path)
    print(f"生成完了: {output_path}")


def create_tx101_lp_capture():
    """TX-101: LP画面キャプチャ報告書（PDF）"""
    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page { size: A4; margin: 20mm; }
            body { font-family: 'Noto Sans CJK JP', sans-serif; font-size: 10pt; line-height: 1.6; }
            .title { text-align: center; font-size: 18pt; font-weight: bold; margin-bottom: 20px; }
            .info-box { background: #f5f5f5; padding: 15px; margin: 15px 0; border-radius: 5px; }
            .screenshot { border: 1px solid #ddd; padding: 20px; margin: 20px 0; background: #fafafa; text-align: center; }
            .screenshot-placeholder { background: #e0e0e0; height: 300px; display: flex; align-items: center; justify-content: center; color: #666; }
            table { width: 100%; border-collapse: collapse; margin: 15px 0; }
            th, td { padding: 8px; border: 1px solid #ddd; }
            th { background: #1F4E79; color: white; }
        </style>
    </head>
    <body>
        <div class="title">ランディングページ制作 納品報告書</div>
        
        <div class="info-box">
            <strong>取引ID:</strong> TX-101<br>
            <strong>件名:</strong> ランディングページ制作<br>
            <strong>納品日:</strong> 2025年9月30日<br>
            <strong>制作会社:</strong> 株式会社クリエイティブワークス
        </div>
        
        <h2 style="color: #1F4E79;">1. 納品物概要</h2>
        <table>
            <tr><th>項目</th><th>内容</th></tr>
            <tr><td>ページ数</td><td>1ページ（縦長LP）</td></tr>
            <tr><td>レスポンシブ対応</td><td>PC / タブレット / スマートフォン</td></tr>
            <tr><td>フォーム</td><td>お問い合わせフォーム設置済み</td></tr>
            <tr><td>URL</td><td>https://example.com/lp/campaign2025/</td></tr>
        </table>
        
        <h2 style="color: #1F4E79;">2. 画面構成</h2>
        
        <h3>ファーストビュー</h3>
        <div class="screenshot">
            <div class="screenshot-placeholder">
                [ファーストビュー画面キャプチャ]<br>
                キャッチコピー + メインビジュアル + CTAボタン
            </div>
        </div>
        
        <h3>サービス紹介セクション</h3>
        <div class="screenshot">
            <div class="screenshot-placeholder">
                [サービス紹介セクション画面キャプチャ]<br>
                3カラムレイアウト + アイコン + 説明文
            </div>
        </div>
        
        <h3>お客様の声セクション</h3>
        <div class="screenshot">
            <div class="screenshot-placeholder">
                [お客様の声セクション画面キャプチャ]<br>
                カルーセル形式 + 顔写真 + コメント
            </div>
        </div>
        
        <h3>お問い合わせフォーム</h3>
        <div class="screenshot">
            <div class="screenshot-placeholder">
                [フォームセクション画面キャプチャ]<br>
                入力フォーム + 送信ボタン + プライバシーポリシー
            </div>
        </div>
        
        <h2 style="color: #1F4E79;">3. 技術仕様</h2>
        <table>
            <tr><th>項目</th><th>仕様</th></tr>
            <tr><td>HTML</td><td>HTML5</td></tr>
            <tr><td>CSS</td><td>CSS3 + Flexbox</td></tr>
            <tr><td>JavaScript</td><td>jQuery 3.6.0</td></tr>
            <tr><td>フォーム</td><td>PHP + メール送信機能</td></tr>
            <tr><td>対応ブラウザ</td><td>Chrome, Safari, Firefox, Edge (最新版)</td></tr>
        </table>
        
        <div class="info-box" style="margin-top: 30px;">
            <strong>備考:</strong><br>
            修正対応は納品後2週間以内、軽微な修正2回まで無償対応いたします。
        </div>
    </body>
    </html>
    """
    output_path = '/home/ubuntu/ad_audit_test_data/検収資料/PDF/TX-101_LP納品報告書.pdf'
    HTML(string=html).write_pdf(output_path)
    print(f"生成完了: {output_path}")


def create_tx102_ad_report():
    """TX-102: リスティング広告運用レポート（Excel）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "運用サマリー"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    
    ws['B2'] = "リスティング広告運用レポート"
    ws['B2'].font = Font(name='Source Serif Pro', size=18, bold=True, color=THEME['primary'])
    
    ws['B3'] = "期間: 2025年10月〜12月"
    ws['B3'].font = Font(name='Source Sans Pro', size=11, color='666666')
    
    ws['B4'] = "運用担当: 個人事業主 山田太郎"
    ws['B4'].font = Font(name='Source Sans Pro', size=11, color='666666')
    
    header_fill = PatternFill(start_color=THEME['primary'], end_color=THEME['primary'], fill_type='solid')
    header_font = Font(name='Source Serif Pro', size=10, bold=True, color='FFFFFF')
    data_font = Font(name='Source Sans Pro', size=10)
    
    # Google広告
    ws['B6'] = "Google広告 運用実績"
    ws['B6'].font = Font(name='Source Serif Pro', size=14, bold=True, color=THEME['primary'])
    
    headers = ["指標", "10月", "11月", "12月", "合計"]
    for col, h in enumerate(headers, start=2):
        cell = ws.cell(row=8, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    google_data = [
        ["インプレッション", 125000, 142000, 138000, 405000],
        ["クリック数", 3750, 4260, 4140, 12150],
        ["CTR", "3.0%", "3.0%", "3.0%", "3.0%"],
        ["CPC", "¥85", "¥82", "¥80", "¥82"],
        ["広告費", "¥318,750", "¥349,320", "¥331,200", "¥999,270"],
        ["CV数", 45, 52, 48, 145],
        ["CVR", "1.2%", "1.2%", "1.2%", "1.2%"],
        ["CPA", "¥7,083", "¥6,718", "¥6,900", "¥6,891"],
    ]
    
    for row_idx, row_data in enumerate(google_data, start=9):
        for col_idx, value in enumerate(row_data, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='right' if col_idx > 2 else 'left', vertical='center')
    
    apply_borders(ws, 8, 16, 2, 6)
    
    # Yahoo!広告
    ws['B19'] = "Yahoo!広告 運用実績"
    ws['B19'].font = Font(name='Source Serif Pro', size=14, bold=True, color=THEME['primary'])
    
    for col, h in enumerate(headers, start=2):
        cell = ws.cell(row=21, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    yahoo_data = [
        ["インプレッション", 85000, 92000, 88000, 265000],
        ["クリック数", 2125, 2300, 2200, 6625],
        ["CTR", "2.5%", "2.5%", "2.5%", "2.5%"],
        ["CPC", "¥75", "¥72", "¥70", "¥72"],
        ["広告費", "¥159,375", "¥165,600", "¥154,000", "¥478,975"],
        ["CV数", 21, 25, 23, 69],
        ["CVR", "1.0%", "1.1%", "1.0%", "1.0%"],
        ["CPA", "¥7,589", "¥6,624", "¥6,696", "¥6,942"],
    ]
    
    for row_idx, row_data in enumerate(yahoo_data, start=22):
        for col_idx, value in enumerate(row_data, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='right' if col_idx > 2 else 'left', vertical='center')
    
    apply_borders(ws, 21, 29, 2, 6)
    
    for i, w in enumerate([20, 15, 15, 15, 15], start=2):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    output_path = '/home/ubuntu/ad_audit_test_data/検収資料/Excel/TX-102_広告運用レポート.xlsx'
    wb.save(output_path)
    print(f"生成完了: {output_path}")


def create_tx103_market_report():
    """TX-103: 市場調査レポート（PDF）- 発注先属性不整合"""
    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page { size: A4; margin: 20mm; }
            body { font-family: 'Noto Sans CJK JP', sans-serif; font-size: 10pt; line-height: 1.6; }
            .title { text-align: center; font-size: 20pt; font-weight: bold; margin-bottom: 10px; color: #1F4E79; }
            .subtitle { text-align: center; font-size: 12pt; color: #666; margin-bottom: 30px; }
            .section { margin: 25px 0; }
            .section h2 { color: #1F4E79; border-bottom: 2px solid #1F4E79; padding-bottom: 5px; }
            table { width: 100%; border-collapse: collapse; margin: 15px 0; }
            th, td { padding: 10px; border: 1px solid #ddd; }
            th { background: #1F4E79; color: white; }
            .highlight { background: #fff3cd; padding: 15px; border-left: 4px solid #ffc107; margin: 15px 0; }
            .footer { margin-top: 40px; text-align: center; color: #666; font-size: 9pt; }
        </style>
    </head>
    <body>
        <div class="title">デジタル広告市場調査レポート</div>
        <div class="subtitle">2025年度 第3四半期</div>
        
        <div class="section">
            <h2>1. エグゼクティブサマリー</h2>
            <p>本レポートは、デジタル広告市場の動向および競合分析をまとめたものです。
            国内デジタル広告市場は前年比12%成長を記録し、特に動画広告とSNS広告の伸びが顕著です。</p>
            
            <div class="highlight">
                <strong>主要ポイント:</strong>
                <ul>
                    <li>国内デジタル広告市場規模: 約2.8兆円（前年比+12%）</li>
                    <li>動画広告が最も高い成長率（+25%）</li>
                    <li>SNS広告のシェアが拡大（全体の28%）</li>
                </ul>
            </div>
        </div>
        
        <div class="section">
            <h2>2. 市場規模推移</h2>
            <table>
                <tr><th>年度</th><th>市場規模</th><th>前年比</th></tr>
                <tr><td>2022年</td><td>2.2兆円</td><td>+8%</td></tr>
                <tr><td>2023年</td><td>2.5兆円</td><td>+14%</td></tr>
                <tr><td>2024年</td><td>2.8兆円</td><td>+12%</td></tr>
                <tr><td>2025年（予測）</td><td>3.1兆円</td><td>+11%</td></tr>
            </table>
        </div>
        
        <div class="section">
            <h2>3. 広告種別シェア</h2>
            <table>
                <tr><th>広告種別</th><th>シェア</th><th>前年比成長率</th></tr>
                <tr><td>検索連動型広告</td><td>35%</td><td>+8%</td></tr>
                <tr><td>ディスプレイ広告</td><td>22%</td><td>+5%</td></tr>
                <tr><td>動画広告</td><td>18%</td><td>+25%</td></tr>
                <tr><td>SNS広告</td><td>28%</td><td>+18%</td></tr>
                <tr><td>その他</td><td>7%</td><td>+3%</td></tr>
            </table>
        </div>
        
        <div class="section">
            <h2>4. 競合分析</h2>
            <p>主要プレイヤーの動向を分析した結果、以下の傾向が確認されました。</p>
            <table>
                <tr><th>企業</th><th>強み</th><th>注力領域</th></tr>
                <tr><td>A社</td><td>検索広告シェアNo.1</td><td>AI活用の自動最適化</td></tr>
                <tr><td>B社</td><td>SNS広告の高いリーチ</td><td>動画広告の拡充</td></tr>
                <tr><td>C社</td><td>EC連携の強さ</td><td>リテールメディア</td></tr>
            </table>
        </div>
        
        <div class="section">
            <h2>5. 提言</h2>
            <ul>
                <li>動画広告への投資拡大を推奨</li>
                <li>SNSプラットフォームの多様化対応</li>
                <li>ファーストパーティデータの活用強化</li>
            </ul>
        </div>
        
        <div class="footer">
            <p>作成: 株式会社建設コンサルタント マーケティング調査部</p>
            <p>※本レポートの無断転載・複製を禁じます</p>
        </div>
    </body>
    </html>
    """
    output_path = '/home/ubuntu/ad_audit_test_data/検収資料/PDF/TX-103_市場調査レポート.pdf'
    HTML(string=html).write_pdf(output_path)
    print(f"生成完了: {output_path}")


def create_tx104_article_list():
    """TX-104: SEO記事一覧（Excel）- 金額と内容不整合"""
    wb = Workbook()
    ws = wb.active
    ws.title = "記事一覧"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    
    ws['B2'] = "SEOコンテンツ記事 納品一覧"
    ws['B2'].font = Font(name='Source Serif Pro', size=18, bold=True, color=THEME['primary'])
    
    ws['B3'] = "納品日: 2025年8月31日 / 納品数: 50本"
    ws['B3'].font = Font(name='Source Sans Pro', size=11, color='666666')
    
    header_fill = PatternFill(start_color=THEME['primary'], end_color=THEME['primary'], fill_type='solid')
    header_font = Font(name='Source Serif Pro', size=10, bold=True, color='FFFFFF')
    data_font = Font(name='Source Sans Pro', size=10)
    
    headers = ["No.", "記事タイトル", "メインKW", "文字数", "ステータス"]
    for col, h in enumerate(headers, start=2):
        cell = ws.cell(row=5, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 50本の記事データ（一部抜粋として表示）
    articles = [
        [1, "デジタルマーケティングとは？基礎から学ぶ入門ガイド", "デジタルマーケティング", 3200, "納品済"],
        [2, "SEO対策の基本｜初心者でもわかる検索上位表示のコツ", "SEO対策 基本", 3150, "納品済"],
        [3, "コンテンツマーケティング成功事例10選", "コンテンツマーケティング 事例", 3400, "納品済"],
        [4, "SNSマーケティング戦略の立て方完全ガイド", "SNSマーケティング", 3100, "納品済"],
        [5, "リスティング広告の費用対効果を最大化する方法", "リスティング広告 費用", 3250, "納品済"],
        [6, "動画マーケティングのトレンドと活用法", "動画マーケティング", 3050, "納品済"],
        [7, "メールマーケティングで成果を出すポイント", "メールマーケティング", 3180, "納品済"],
        [8, "インフルエンサーマーケティングの始め方", "インフルエンサーマーケティング", 3300, "納品済"],
        [9, "Web広告の種類と特徴を徹底比較", "Web広告 種類", 3420, "納品済"],
        [10, "ランディングページ最適化（LPO）のベストプラクティス", "LPO", 3150, "納品済"],
    ]
    
    # 残り40本も追加
    topics = [
        "アクセス解析", "コンバージョン最適化", "ブランディング", "カスタマージャーニー",
        "マーケティングオートメーション", "CRM活用", "データドリブンマーケティング",
        "パーソナライゼーション", "オムニチャネル", "D2Cマーケティング",
        "BtoBマーケティング", "リードジェネレーション", "ナーチャリング",
        "ABテスト", "ヒートマップ分析", "ユーザー行動分析", "競合分析",
        "市場調査", "ペルソナ設計", "カスタマーサクセス", "NPS向上",
        "口コミマーケティング", "アフィリエイト", "ネイティブ広告",
        "プログラマティック広告", "DSP活用", "リターゲティング",
        "ダイナミック広告", "ショッピング広告", "アプリマーケティング",
        "ASO対策", "プッシュ通知", "チャットボット", "音声検索対策",
        "ローカルSEO", "MEO対策", "構造化データ", "Core Web Vitals",
        "モバイルファースト", "AMP対応"
    ]
    
    for i, topic in enumerate(topics, start=11):
        articles.append([i, f"{topic}の基礎知識と実践テクニック", topic, 3000 + (i * 5) % 500, "納品済"])
    
    for row_idx, row_data in enumerate(articles, start=6):
        for col_idx, value in enumerate(row_data, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            if col_idx == 2:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx == 5:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    
    apply_borders(ws, 5, 5 + len(articles), 2, 6)
    
    for i, w in enumerate([6, 50, 25, 10, 10], start=2):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    output_path = '/home/ubuntu/ad_audit_test_data/検収資料/Excel/TX-104_SEO記事一覧.xlsx'
    wb.save(output_path)
    print(f"生成完了: {output_path}")


def create_tx106_influencer_list():
    """TX-106: インフルエンサー投稿一覧（Excel）- 数量不整合"""
    wb = Workbook()
    ws = wb.active
    ws.title = "投稿一覧"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    
    ws['B2'] = "インフルエンサー施策 投稿一覧"
    ws['B2'].font = Font(name='Source Serif Pro', size=18, bold=True, color=THEME['primary'])
    
    ws['B3'] = "キャンペーン期間: 2025年8月〜9月"
    ws['B3'].font = Font(name='Source Sans Pro', size=11, color='666666')
    
    # 注意書き（数量不整合を示唆）
    ws['B4'] = "※発注: 10名 → 実施: 7名（3名はスケジュール都合により未実施）"
    ws['B4'].font = Font(name='Source Sans Pro', size=10, color='C62828', italic=True)
    
    header_fill = PatternFill(start_color=THEME['primary'], end_color=THEME['primary'], fill_type='solid')
    header_font = Font(name='Source Serif Pro', size=10, bold=True, color='FFFFFF')
    data_font = Font(name='Source Sans Pro', size=10)
    
    headers = ["No.", "インフルエンサー名", "フォロワー数", "投稿日", "いいね", "コメント", "リーチ", "ステータス"]
    for col, h in enumerate(headers, start=2):
        cell = ws.cell(row=6, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 7名分のデータ（10名発注だが7名のみ実施）
    influencers = [
        [1, "@beauty_mika", "152,000", "2025/08/15", 4520, 128, 45000, "完了"],
        [2, "@lifestyle_yuki", "198,000", "2025/08/18", 5890, 156, 58000, "完了"],
        [3, "@fashion_rina", "175,000", "2025/08/22", 4980, 142, 52000, "完了"],
        [4, "@travel_ken", "142,000", "2025/09/01", 3850, 98, 38000, "完了"],
        [5, "@food_hana", "168,000", "2025/09/05", 5120, 167, 49000, "完了"],
        [6, "@fitness_taro", "135,000", "2025/09/12", 3680, 89, 35000, "完了"],
        [7, "@tech_sato", "188,000", "2025/09/20", 4750, 134, 55000, "完了"],
    ]
    
    for row_idx, row_data in enumerate(influencers, start=7):
        for col_idx, value in enumerate(row_data, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            if col_idx in [6, 7, 8]:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if isinstance(value, int):
                    cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal='left' if col_idx > 2 else 'center', vertical='center')
    
    apply_borders(ws, 6, 13, 2, 9)
    
    # サマリー
    ws['B16'] = "実績サマリー"
    ws['B16'].font = Font(name='Source Serif Pro', size=14, bold=True, color=THEME['primary'])
    
    summary = [
        ["発注人数", "10名"],
        ["実施人数", "7名"],
        ["未実施人数", "3名"],
        ["総いいね数", "32,790"],
        ["総コメント数", "914"],
        ["総リーチ数", "332,000"],
        ["請求金額", "¥2,100,000（7名×¥300,000）"],
    ]
    
    for row_idx, (label, value) in enumerate(summary, start=17):
        ws.cell(row=row_idx, column=2, value=label).font = data_font
        cell = ws.cell(row=row_idx, column=3, value=value)
        cell.font = Font(name='Source Sans Pro', size=10, bold=True)
    
    for i, w in enumerate([6, 20, 12, 12, 10, 10, 10, 10], start=2):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    output_path = '/home/ubuntu/ad_audit_test_data/検収資料/Excel/TX-106_インフルエンサー投稿一覧.xlsx'
    wb.save(output_path)
    print(f"生成完了: {output_path}")


def create_tx107_system_spec():
    """TX-107: システム仕様書（PDF）- 日付不整合"""
    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page { size: A4; margin: 20mm; }
            body { font-family: 'Noto Sans CJK JP', sans-serif; font-size: 10pt; line-height: 1.6; }
            .title { text-align: center; font-size: 20pt; font-weight: bold; color: #1F4E79; margin-bottom: 5px; }
            .subtitle { text-align: center; font-size: 12pt; color: #666; margin-bottom: 30px; }
            .section { margin: 20px 0; }
            .section h2 { color: #1F4E79; border-left: 4px solid #1F4E79; padding-left: 10px; }
            table { width: 100%; border-collapse: collapse; margin: 15px 0; }
            th, td { padding: 8px; border: 1px solid #ddd; }
            th { background: #1F4E79; color: white; }
            .code { background: #f5f5f5; padding: 10px; font-family: monospace; font-size: 9pt; overflow-x: auto; }
            .warning { background: #fff3cd; padding: 10px; border-left: 4px solid #ffc107; margin: 10px 0; }
        </style>
    </head>
    <body>
        <div class="title">広告効果測定ダッシュボード</div>
        <div class="subtitle">システム仕様書 Ver.1.0</div>
        
        <div class="section">
            <h2>1. システム概要</h2>
            <table>
                <tr><th width="30%">項目</th><th>内容</th></tr>
                <tr><td>システム名</td><td>広告効果測定ダッシュボード</td></tr>
                <tr><td>バージョン</td><td>1.0.0</td></tr>
                <tr><td>開発会社</td><td>システム開発株式会社テクノソリューション</td></tr>
                <tr><td>納品日</td><td>2025年10月15日</td></tr>
            </table>
        </div>
        
        <div class="section">
            <h2>2. 機能一覧</h2>
            <table>
                <tr><th>機能ID</th><th>機能名</th><th>概要</th></tr>
                <tr><td>F001</td><td>データ連携</td><td>Google/Yahoo!/Meta広告データの自動取得</td></tr>
                <tr><td>F002</td><td>ダッシュボード</td><td>KPI可視化、グラフ表示</td></tr>
                <tr><td>F003</td><td>レポート出力</td><td>PDF/Excel形式でのレポート自動生成</td></tr>
                <tr><td>F004</td><td>アラート機能</td><td>閾値超過時のメール通知</td></tr>
                <tr><td>F005</td><td>ユーザー管理</td><td>権限別アクセス制御</td></tr>
            </table>
        </div>
        
        <div class="section">
            <h2>3. システム構成</h2>
            <table>
                <tr><th>コンポーネント</th><th>技術スタック</th></tr>
                <tr><td>フロントエンド</td><td>React 18.x + TypeScript</td></tr>
                <tr><td>バックエンド</td><td>Python 3.11 + FastAPI</td></tr>
                <tr><td>データベース</td><td>PostgreSQL 15</td></tr>
                <tr><td>キャッシュ</td><td>Redis 7.x</td></tr>
                <tr><td>インフラ</td><td>AWS (ECS, RDS, ElastiCache)</td></tr>
            </table>
        </div>
        
        <div class="section">
            <h2>4. API仕様（抜粋）</h2>
            <h3>4.1 広告データ取得API</h3>
            <div class="code">
GET /api/v1/ads/performance
Query Parameters:
  - start_date: string (YYYY-MM-DD)
  - end_date: string (YYYY-MM-DD)
  - platform: string (google|yahoo|meta)
  
Response:
{
  "data": [
    {
      "date": "2025-10-01",
      "impressions": 125000,
      "clicks": 3750,
      "cost": 318750,
      "conversions": 45
    }
  ]
}
            </div>
        </div>
        
        <div class="section">
            <h2>5. 画面一覧</h2>
            <table>
                <tr><th>画面ID</th><th>画面名</th><th>概要</th></tr>
                <tr><td>SCR001</td><td>ログイン</td><td>認証画面</td></tr>
                <tr><td>SCR002</td><td>ダッシュボード</td><td>KPIサマリー表示</td></tr>
                <tr><td>SCR003</td><td>詳細分析</td><td>媒体別詳細データ</td></tr>
                <tr><td>SCR004</td><td>レポート</td><td>レポート生成・出力</td></tr>
                <tr><td>SCR005</td><td>設定</td><td>アラート・ユーザー設定</td></tr>
            </table>
        </div>
        
        <div class="section">
            <h2>6. 非機能要件</h2>
            <table>
                <tr><th>項目</th><th>要件</th></tr>
                <tr><td>可用性</td><td>99.5%以上</td></tr>
                <tr><td>応答時間</td><td>画面表示3秒以内</td></tr>
                <tr><td>同時接続数</td><td>100ユーザー</td></tr>
                <tr><td>データ保持期間</td><td>3年間</td></tr>
            </table>
        </div>
        
        <div class="warning">
            <strong>注意:</strong> 本仕様書は納品時点の内容です。
            運用開始後の仕様変更は別途ご相談ください。
        </div>
    </body>
    </html>
    """
    output_path = '/home/ubuntu/ad_audit_test_data/検収資料/PDF/TX-107_システム仕様書.pdf'
    HTML(string=html).write_pdf(output_path)
    print(f"生成完了: {output_path}")


if __name__ == "__main__":
    # 各検収資料を生成
    create_tx001_banner_report()
    create_tx002_sns_report()
    create_tx003_video_spec()
    create_tx101_lp_capture()
    create_tx102_ad_report()
    create_tx103_market_report()
    create_tx104_article_list()
    create_tx106_influencer_list()
    create_tx107_system_spec()
    
    print("\n検収資料（PDF、Excel）の生成が完了しました。")
